from openpyxl.styles import Border, Side, Font, GradientFill, Alignment
import datetime
from openpyxl import Workbook

wo = "7564"
wb = Workbook()

ws = wb.create_sheet(f"Int.mod.{wo}")
ws.title = wo

def sheet_generate(wo, ws, customer, qty, date_start, date_required, part_description, part_no, drawing_no, issue_no, contact, bin, part_to_issue):

    border_ranges = []
    
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def border_setting(ws, border_ranges):
        for range in border_ranges:
            set_border(ws, range)

    

wb.remove(wb['Sheet'])
wb.save(filename=f"IntMod{wo}.xlsx")